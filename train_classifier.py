"""
Training script for CompressedEmbeddingsClassifier.

This script:
1. Reads PEP_MIMIC_SPR_DATA.xlsx
2. Extracts PDB codes and downloads FASTA files
3. Identifies target chains (longer sequence) and binder chains
4. Creates a dataset with target sequence, peptide sequence, and binary label
5. Trains the classifier with 80-20 train/test split
"""

import sys
import os
from pathlib import Path
import re
import urllib.request
import time
from typing import List, Tuple, Optional
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
import numpy as np
from tqdm import tqdm

# Try to import pandas/openpyxl, with fallback
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("Warning: pandas not available. Trying alternative methods...")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from compressed_embeddings import CompressedEmbeddingsClassifier


class SequencePairDataset(Dataset):
    """Dataset for sequence pairs."""
    
    def __init__(self, target_sequences: List[str], peptide_sequences: List[str], labels: List[int]):
        self.target_sequences = target_sequences
        self.peptide_sequences = peptide_sequences
        self.labels = labels
        
    def __len__(self):
        return len(self.labels)
    
    def __getitem__(self, idx):
        return {
            'target_seq': self.target_sequences[idx],
            'peptide_seq': self.peptide_sequences[idx],
            'label': self.labels[idx]
        }


def read_excel_file(filepath: str) -> List[dict]:
    """Read Excel file using available methods."""
    if HAS_PANDAS:
        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            return df.to_dict('records')
        except Exception as e:
            print(f"Error reading with pandas: {e}")
    
    # Fallback: try openpyxl directly
    if HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Read data
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(dict(zip(headers, row)))
            return data
        except Exception as e:
            print(f"Error reading with openpyxl: {e}")
    
    raise RuntimeError("Could not read Excel file. Please install pandas and openpyxl.")


def download_fasta(pdb_code: str, output_dir: Path) -> Optional[str]:
    """Download FASTA file from PDB for a given PDB code."""
    pdb_code = pdb_code.upper()
    fasta_path = output_dir / f"{pdb_code}.fasta"
    
    # If already downloaded, return path
    if fasta_path.exists():
        return str(fasta_path)
    
    # Try to download from PDB
    url = f"https://www.rcsb.org/fasta/entry/{pdb_code}/display"
    
    try:
        urllib.request.urlretrieve(url, fasta_path)
        time.sleep(0.1)  # Be nice to the server
        return str(fasta_path)
    except Exception as e:
        print(f"Warning: Could not download FASTA for {pdb_code}: {e}")
        return None


def parse_fasta(fasta_path: str) -> List[Tuple[str, str]]:
    """Parse FASTA file and return list of (header, sequence) tuples."""
    sequences = []
    current_header = None
    current_sequence = []
    
    with open(fasta_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('>'):
                if current_header is not None:
                    sequences.append((current_header, ''.join(current_sequence)))
                current_header = line[1:]  # Remove '>'
                current_sequence = []
            else:
                current_sequence.append(line)
        
        # Add last sequence
        if current_header is not None:
            sequences.append((current_header, ''.join(current_sequence)))
    
    return sequences


def get_target_chain_sequence(pdb_code: str, fasta_dir: Path) -> Optional[str]:
    """Get the target chain sequence (longer of the two chains)."""
    fasta_path = download_fasta(pdb_code, fasta_dir)
    if fasta_path is None:
        return None
    
    sequences = parse_fasta(fasta_path)
    
    if len(sequences) < 2:
        print(f"Warning: {pdb_code} has fewer than 2 chains")
        return None
    
    # Find the longer sequence (target chain)
    sequences_with_lengths = [(header, seq, len(seq)) for header, seq in sequences]
    sequences_with_lengths.sort(key=lambda x: x[2], reverse=True)
    
    # Return the longest sequence
    return sequences_with_lengths[0][1]


def process_data(excel_path: str, output_txt_path: str, fasta_dir: Path):
    """Process Excel file and create training data text file."""
    print("Reading Excel file...")
    data = read_excel_file(excel_path)
    
    if not data:
        raise ValueError("No data found in Excel file")
    
    # Print column names for debugging
    print(f"Found {len(data)} rows")
    if data:
        print(f"Columns: {list(data[0].keys())}")
    
    # Find column indices
    # First column should be name, second should be peptide sequence
    # Need to find KD column
    first_key = list(data[0].keys())[0]
    second_key = list(data[0].keys())[1] if len(data[0].keys()) > 1 else None
    
    # Find KD column (look for columns with 'KD' or 'kd' in name)
    kd_key = None
    for key in data[0].keys():
        if 'kd' in str(key).lower():
            kd_key = key
            break
    
    print(f"Using columns: name={first_key}, peptide={second_key}, KD={kd_key}")
    
    # Create output directory for FASTA files
    fasta_dir.mkdir(exist_ok=True)
    
    results = []
    
    print("\nProcessing data...")
    for i, row in enumerate(tqdm(data, desc="Processing rows")):
        try:
            # Get name (first column)
            name = str(row.get(first_key, ''))
            if not name or name == 'nan':
                continue
            
            # Extract PDB code (part before first underscore)
            pdb_code = name.split('_')[0].strip()
            if not pdb_code or len(pdb_code) < 4:
                continue
            
            # Get peptide sequence (second column)
            peptide_seq = str(row.get(second_key, '')).strip()
            if not peptide_seq or peptide_seq == 'nan' or len(peptide_seq) < 3:
                continue
            
            # Get target chain sequence
            target_seq = get_target_chain_sequence(pdb_code, fasta_dir)
            if target_seq is None:
                continue
            
            # Get label based on KD column
            kd_value = row.get(kd_key, None)
            if kd_value is None:
                label = 0
            elif HAS_PANDAS and pd.isna(kd_value):
                label = 0
            elif kd_value == '' or str(kd_value).lower() == 'nan':
                label = 0
            else:
                # Try to convert to number
                try:
                    float(str(kd_value))
                    label = 1
                except (ValueError, TypeError):
                    label = 0
            
            results.append({
                'target_seq': target_seq,
                'peptide_seq': peptide_seq,
                'label': label
            })
            
        except Exception as e:
            print(f"\nError processing row {i}: {e}")
            continue
    
    print(f"\nSuccessfully processed {len(results)} samples")
    
    # Write to text file
    print(f"\nWriting results to {output_txt_path}...")
    with open(output_txt_path, 'w') as f:
        f.write("target_sequence\tpeptide_sequence\tlabel\n")
        for result in results:
            f.write(f"{result['target_seq']}\t{result['peptide_seq']}\t{result['label']}\n")
    
    print(f"Created dataset file: {output_txt_path}")
    return results


def load_dataset_from_txt(txt_path: str) -> Tuple[List[str], List[str], List[int]]:
    """Load dataset from text file."""
    target_sequences = []
    peptide_sequences = []
    labels = []
    
    with open(txt_path, 'r') as f:
        # Skip header
        next(f)
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split('\t')
            if len(parts) >= 3:
                target_sequences.append(parts[0])
                peptide_sequences.append(parts[1])
                labels.append(int(parts[2]))
    
    return target_sequences, peptide_sequences, labels


def train_classifier(
    train_dataset: SequencePairDataset,
    test_dataset: SequencePairDataset,
    model: CompressedEmbeddingsClassifier,
    num_epochs: int = 10,
    batch_size: int = 8,
    learning_rate: float = 1e-4,
    device: str = "cuda"
):
    """Train the classifier."""
    
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False)
    
    # Loss and optimizer
    # Only optimize MLP parameters (encoder is frozen)
    criterion = nn.CrossEntropyLoss()
    trainable_params = [p for p in model.parameters() if p.requires_grad]
    optimizer = optim.Adam(trainable_params, lr=learning_rate)
    print(f"Training {sum(p.numel() for p in trainable_params)} parameters")
    
    # Training loop
    model.train()
    for epoch in range(num_epochs):
        total_loss = 0
        correct = 0
        total = 0
        
        print(f"\nEpoch {epoch + 1}/{num_epochs}")
        for batch in tqdm(train_loader, desc="Training"):
            # Convert batch to lists (DataLoader returns lists for string data)
            target_seqs = batch['target_seq'] if isinstance(batch['target_seq'], list) else list(batch['target_seq'])
            peptide_seqs = batch['peptide_seq'] if isinstance(batch['peptide_seq'], list) else list(batch['peptide_seq'])
            labels = torch.tensor(batch['label'], dtype=torch.long).to(device)
            
            optimizer.zero_grad()
            
            # Forward pass
            logits = model(target_seqs, peptide_seqs)
            loss = criterion(logits, labels)
            
            # Backward pass
            loss.backward()
            optimizer.step()
            
            total_loss += loss.item()
            _, predicted = torch.max(logits.data, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()
        
        train_acc = 100 * correct / total
        avg_loss = total_loss / len(train_loader)
        
        print(f"Train Loss: {avg_loss:.4f}, Train Acc: {train_acc:.2f}%")
        
        # Evaluate on test set
        model.eval()
        test_correct = 0
        test_total = 0
        with torch.no_grad():
            for batch in test_loader:
                # Convert batch to lists
                target_seqs = batch['target_seq'] if isinstance(batch['target_seq'], list) else list(batch['target_seq'])
                peptide_seqs = batch['peptide_seq'] if isinstance(batch['peptide_seq'], list) else list(batch['peptide_seq'])
                labels = torch.tensor(batch['label'], dtype=torch.long).to(device)
                
                logits = model(target_seqs, peptide_seqs)
                _, predicted = torch.max(logits.data, 1)
                test_total += labels.size(0)
                test_correct += (predicted == labels).sum().item()
        
        test_acc = 100 * test_correct / test_total
        print(f"Test Acc: {test_acc:.2f}%")
        model.train()
    
    return model


def main():
    """Main training function."""
    # Paths
    excel_path = "PEP_MIMIC_SPR_DATA.xlsx"
    output_txt_path = "training_data.txt"
    fasta_dir = Path("fasta_files")
    model_save_path = "trained_classifier.pt"
    
    # Check if Excel file exists
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Process data if text file doesn't exist
    if not os.path.exists(output_txt_path):
        print("Processing data from Excel file...")
        process_data(excel_path, output_txt_path, fasta_dir)
    else:
        print(f"Using existing dataset file: {output_txt_path}")
    
    # Load dataset
    print("\nLoading dataset...")
    target_sequences, peptide_sequences, labels = load_dataset_from_txt(output_txt_path)
    
    print(f"Loaded {len(target_sequences)} samples")
    print(f"Positive samples: {sum(labels)}, Negative samples: {len(labels) - sum(labels)}")
    
    # Split into train/test (80-20)
    split_idx = int(0.8 * len(target_sequences))
    
    train_target = target_sequences[:split_idx]
    train_peptide = peptide_sequences[:split_idx]
    train_labels = labels[:split_idx]
    
    test_target = target_sequences[split_idx:]
    test_peptide = peptide_sequences[split_idx:]
    test_labels = labels[split_idx:]
    
    print(f"\nTrain set: {len(train_labels)} samples")
    print(f"Test set: {len(test_labels)} samples")
    
    # Create datasets
    train_dataset = SequencePairDataset(train_target, train_peptide, train_labels)
    test_dataset = SequencePairDataset(test_target, test_peptide, test_labels)
    
    # Initialize model
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"\nInitializing model on {device}...")
    
    model = CompressedEmbeddingsClassifier(
        shorten_factor=1,
        dimension=64,
        esm2_model_name="facebook/esm2_t36_3B_UR50D",
        mlp_hidden_dims=[256, 128],
        dropout=0.1,
        device=device
    )
    
    # Make encoder trainable (unfreeze MLP parameters)
    # Note: The encoder's ESM2 and compression model remain frozen,
    # but we can fine-tune the MLP
    for param in model.mlp.parameters():
        param.requires_grad = True
    
    # Train model
    print("\nStarting training...")
    trained_model = train_classifier(
        train_dataset,
        test_dataset,
        model,
        num_epochs=10,
        batch_size=4,  # Smaller batch size due to memory constraints
        learning_rate=1e-4,
        device=device
    )
    
    # Save model
    print(f"\nSaving model to {model_save_path}...")
    torch.save(trained_model.state_dict(), model_save_path)
    print("Training complete!")


if __name__ == "__main__":
    main()

