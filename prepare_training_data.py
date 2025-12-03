"""
Script to prepare training data from Excel file.

This script:
1. Reads PEP_MIMIC_SPR_DATA.xlsx
2. Uses the sheet name as the protein name to search for PDB code
3. Downloads FASTA files from PDB
4. Identifies target chains (longer sequence)
5. Creates a dataset file with target sequence, peptide sequence, and binary label
"""

import sys
import os
from pathlib import Path
import re
import urllib.request
import urllib.parse
import json
import time
import ssl
from typing import List, Tuple, Optional
from tqdm import tqdm

# Try to import pandas/openpyxl, with fallback
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("Warning: pandas not available. Trying alternative methods...")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Create unverified SSL context to bypass certificate errors
ssl._create_default_https_context = ssl._create_unverified_context


def read_excel_file(filepath: str) -> Tuple[List[dict], str]:
    """Read Excel file using available methods.
    
    Returns:
        Tuple of (data, sheet_name) where sheet_name is the name of the active sheet
    """
    if HAS_PANDAS:
        try:
            # Read the first sheet and get its name
            xl_file = pd.ExcelFile(filepath, engine='openpyxl')
            sheet_name = xl_file.sheet_names[0]
            df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')
            return df.to_dict('records'), sheet_name
        except Exception as e:
            print(f"Error reading with pandas: {e}")
    
    # Fallback: try openpyxl directly
    if HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            sheet_name = ws.title
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Read data
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(dict(zip(headers, row)))
            return data, sheet_name
        except Exception as e:
            print(f"Error reading with openpyxl: {e}")
    
    raise RuntimeError("Could not read Excel file. Please install pandas and openpyxl.")


def search_pdb_by_name(protein_name: str) -> Optional[str]:
    """Search PDB for a protein name and return the best matching PDB code."""
    print(f"Searching PDB for: {protein_name}")
    
    # Clean name (remove non-alphanumeric except spaces)
    clean_name = re.sub(r'[^a-zA-Z0-9\s]', ' ', protein_name).strip()
    if not clean_name:
        return None
        
    url = "https://search.rcsb.org/rcsbsearch/v2/query"
    
    # RCSB Search API query
    # Simplified query structure
    query = {
      "query": {
        "type": "terminal",
        "service": "text",
        "parameters": {
          "value": clean_name
        }
      },
      "return_type": "entry",
      "request_options": {
        "return_all_hits": True
      }
    }
    
    try:
        data = json.dumps(query).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers={'Content-Type': 'application/json'})
        
        with urllib.request.urlopen(req) as response:
            if response.status == 200:
                result = json.loads(response.read().decode('utf-8'))
                if result.get('result_set'):
                    # Return the first hit
                    pdb_code = result['result_set'][0]['identifier']
                    print(f"Found PDB code: {pdb_code}")
                    return pdb_code
                else:
                    print("No results found in PDB.")
            else:
                print(f"Search failed with status: {response.status}")
                
    except Exception as e:
        print(f"Error searching PDB: {e}")
        
    return None


def download_fasta(pdb_code: str, output_dir: Path) -> Optional[str]:
    """Download FASTA file from PDB for a given PDB code."""
    pdb_code = pdb_code.upper().strip()
    fasta_path = output_dir / f"{pdb_code}.fasta"
    
    # If already downloaded, return path
    if fasta_path.exists() and fasta_path.stat().st_size > 100:
        return str(fasta_path)
    
    # Try to download from PDB
    url = f"https://www.rcsb.org/fasta/entry/{pdb_code}/display"
    
    print(f"Downloading FASTA for {pdb_code} from {url}...")
    try:
        # Use custom opener to bypass SSL verification
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        with urllib.request.urlopen(url, context=ctx) as response, open(fasta_path, 'wb') as out_file:
            out_file.write(response.read())
            
        time.sleep(0.1)  # Be nice to the server
        
        # Check if the file was actually downloaded (not an error page)
        if fasta_path.exists() and fasta_path.stat().st_size > 100:
            return str(fasta_path)
        else:
            # Delete if it's too small (likely an error page)
            if fasta_path.exists():
                fasta_path.unlink()
            print(f"Download failed (file too small or empty)")
            return None
    except Exception as e:
        print(f"Warning: Could not download FASTA for {pdb_code}: {e}")
        return None


def parse_fasta(fasta_path: str) -> List[Tuple[str, str]]:
    """Parse FASTA file and return list of (header, sequence) tuples."""
    sequences = []
    current_header = None
    current_sequence = []
    
    with open(fasta_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line.startswith('>'):
                if current_header is not None:
                    sequences.append((current_header, ''.join(current_sequence)))
                current_header = line[1:]  # Remove '>'
                current_sequence = []
            else:
                current_sequence.append(line)
        
        # Add last sequence
        if current_header is not None:
            sequences.append((current_header, ''.join(current_sequence)))
    
    return sequences


def get_target_chain_sequence(pdb_code: str, fasta_dir: Path) -> Optional[str]:
    """Get the target chain sequence (longer of the two chains)."""
    fasta_path = download_fasta(pdb_code, fasta_dir)
    if fasta_path is None:
        return None
    
    sequences = parse_fasta(fasta_path)
    
    if len(sequences) < 2:
        print(f"Warning: {pdb_code} has fewer than 2 chains. Proceeding anyway using the longest chain.")
    
    # Find the longer sequence (target chain)
    sequences_with_lengths = [(header, seq, len(seq)) for header, seq in sequences]
    sequences_with_lengths.sort(key=lambda x: x[2], reverse=True)
    
    # Return the longest sequence
    return sequences_with_lengths[0][1]


def main():
    """Main function to process data."""
    # Paths
    excel_path = "PEP_MIMIC_SPR_DATA.xlsx"
    output_txt_path = "training_data.txt"
    fasta_dir = Path("fasta_files")
    
    # Check if Excel file exists
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    print("Reading Excel file...")
    data, sheet_name = read_excel_file(excel_path)
    
    if not data:
        raise ValueError("No data found in Excel file")
    
    # Extract PDB code from sheet name
    # Use sheet name as search term if it's not a PDB code
    raw_name = sheet_name.strip()
    
    # Check if it looks like a PDB code (4 chars)
    if len(raw_name) == 4 and re.match(r'^[0-9][A-Za-z0-9]{3}$', raw_name):
        pdb_code = raw_name.upper()
        print(f"Sheet name '{raw_name}' looks like a PDB code. Using it directly.")
    else:
        # Not a PDB code, assume it's a protein name
        # If it contains special chars like underscores, clean it up
        search_term = raw_name
        if '-' in search_term:
            search_term = search_term.split('-')[0]
        if '_' in search_term:
            search_term = search_term.split('_')[0]
            
        print(f"Sheet name '{raw_name}' is not a PDB code. Searching for '{search_term}'...")
        pdb_code = search_pdb_by_name(search_term)
        
        if not pdb_code:
            # Fallback: try searching for the full name
            print(f"Retrying search with full name '{raw_name}'...")
            pdb_code = search_pdb_by_name(raw_name)
            
        if not pdb_code:
            print("Could not find a PDB code.")
            # Try a known fallback for TROP2 if applicable
            if "TROP2" in raw_name.upper():
                print("Using fallback PDB for TROP2: 7U8I")
                pdb_code = "7U8I"
            else:
                raise ValueError(f"Could not find PDB code for '{raw_name}'")

    print(f"Using PDB code: {pdb_code}")
    
    # Print column names for debugging
    print(f"Found {len(data)} rows")
    if data:
        print(f"Columns: {list(data[0].keys())}")
    
    # Find column indices
    # Second column should be peptide sequence (ignoring first column)
    # Need to find KD column
    column_keys = list(data[0].keys())
    second_key = column_keys[1] if len(column_keys) > 1 else None
    
    # Find KD column (look for columns with 'KD' or 'kd' in name)
    kd_key = None
    for key in data[0].keys():
        if 'kd' in str(key).lower():
            kd_key = key
            break
    
    print(f"Using columns: peptide={second_key}, KD={kd_key}")
    
    # Create output directory for FASTA files
    fasta_dir.mkdir(exist_ok=True)
    
    # Get target chain sequence once (same for all rows since they share the same PDB)
    print(f"\nDownloading FASTA for target protein: {pdb_code}")
    target_seq = get_target_chain_sequence(pdb_code, fasta_dir)
    if target_seq is None:
        raise ValueError(f"Could not get target chain sequence for PDB code: {pdb_code}")
    
    print(f"Target chain sequence length: {len(target_seq)}")
    print(f"Target chain sequence (first 50 aa): {target_seq[:50]}...")
    
    results = []
    
    print("\nProcessing data...")
    for i, row in enumerate(tqdm(data, desc="Processing rows")):
        try:
            # Get peptide sequence (second column, ignoring first column)
            peptide_seq = str(row.get(second_key, '')).strip()
            if not peptide_seq or peptide_seq == 'nan' or len(peptide_seq) < 3:
                continue
            
            # Get label based on KD column
            kd_value = row.get(kd_key, None)
            if kd_value is None:
                label = 0
            elif HAS_PANDAS and pd.isna(kd_value):
                label = 0
            elif kd_value == '' or str(kd_value).lower() == 'nan':
                label = 0
            else:
                # Try to convert to number
                try:
                    float(str(kd_value))
                    label = 1
                except (ValueError, TypeError):
                    label = 0
            
            results.append({
                'target_seq': target_seq,
                'peptide_seq': peptide_seq,
                'label': label
            })
            
        except Exception as e:
            print(f"\nError processing row {i}: {e}")
            continue
    
    print(f"\nSuccessfully processed {len(results)} samples")
    print(f"Positive samples (label=1): {sum(r['label'] for r in results)}")
    print(f"Negative samples (label=0): {len(results) - sum(r['label'] for r in results)}")
    
    # Write to text file
    print(f"\nWriting results to {output_txt_path}...")
    with open(output_txt_path, 'w') as f:
        f.write("target_sequence\tpeptide_sequence\tlabel\n")
        for result in results:
            f.write(f"{result['target_seq']}\t{result['peptide_seq']}\t{result['label']}\n")
    
    print(f"Created dataset file: {output_txt_path}")
    print("\nYou can now check the file before training.")


if __name__ == "__main__":
    main()
